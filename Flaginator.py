from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime  import datetime 
import locale


class App:

    def __init__(self,root):
        self.root = root
        self.root.title("Generator plomby")
        self.root.geometry("500x500")

        self.seal = ""  # Inicjalizacja atrybutu instancji
        self.products = ""
        self.file_path = ""
# sprawdzenie dnia tygodnia
        self.dzis = datetime.today().weekday()
        locale.setlocale(locale.LC_ALL, 'pl_PL.UTF-8')
        self.name_of_day = datetime.today().strftime('%A')

# wczytywanie pliku
        self.btn_load = tk.Button(root, text="Wybierz plik Excel", command=self.select_data)
        self.btn_load.pack(pady=10)

# pokazanie wybranego pliku
        self.seal_label = tk.Label(root, text="Wybrana plomba:")
        self.seal_label.pack(pady=5)

        self.seal_number = tk.Entry(root,width=50, fg="blue")
        self.seal_number.pack(pady=5)
# czy są baterie?

        self.bateries_lit_ion = tk.IntVar()
        self.bateries_lit_met = tk.IntVar()


        self.label_bat = tk.Label(root, text="Załadowane baterie?")
        self.label_bat.pack()
        self.radio_bat2 = tk.Checkbutton(root, text="LIT-ION", variable=self.bateries_lit_ion, onvalue=True, offvalue=False)
        self.radio_bat2.pack()
        self.radio_bat3 = tk.Checkbutton(root, text="LIT-MET", variable=self.bateries_lit_met, onvalue=True, offvalue=False)
        self.radio_bat3.pack()
# Sprawdzenie suchego lodu poprzez zawartość
        self.is_dry_ice = False
        self.is_t09 = False
# typ auta
        self.car_label = tk.Label(root, text="Wybierz typ auta:")
        self.car_label.pack(pady=5)

        self.car_type = tk.IntVar()

        self.r1 = tk.Radiobutton(root, text="COY", variable=self.car_type, value= 1)
        self.r1.pack()
        self.r2 = tk.Radiobutton(root, text="NCY", variable=self.car_type, value= 2)
        self.r2.pack()
        self.r3 = tk.Radiobutton(root, text="CNY", variable=self.car_type, value= 3)
        self.r3.pack()

# dodanie obslugi sobota
        self.saturday = tk.IntVar()

        self.saturday_label = tk.Label(root, text="Czy są paczki na sobotę?")
        self.saturday_label.pack()
        self.checkbox_saturday = tk.Checkbutton(root, text="Tak", variable=self.saturday, onvalue=True, offvalue=False)
        self.checkbox_saturday.pack()
   
# wyliczenie plomby

        self.result_btn = tk.Button(root, text="Stwórz plombe", command=self.check_if_data_available)
        self.result_btn.pack(pady=10)

        self.result_label = tk.Label(root, text="Wygenerowana plomba:")
        self.result_label.pack(pady=5)

        self.result_text = tk.Entry(root,width=50)
        self.result_text.pack(pady=5)


        self.clipboard_button = tk.Button(root, text="Skopiuj do schowka", command=self.copy_to_clipboard)
        self.clipboard_button.pack(pady=10)

    def copy_to_clipboard(self):
        value = self.result_text.get()
        self.root.clipboard_clear()
        self.root.clipboard_append(value)
        self.root.update()

        messagebox.showinfo("Skopiowano", "Plomba została skopiowana do schowka.")

    def read_excel(self, file_path):
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            products = set()
            # znajdz numer kolumny
            header = [cell.value for cell in ws[1]]
            if "Product" not in header:
                raise ValueError("Brak kolumny 'Product' w pliku Excel.")
            
            if "Content" not in header:
                raise ValueError("Brak kolumny 'Content' w pliku Excel.")
            
            product_col = header.index("Product")

            content_col = header.index("Content")

            t09_col = header.index("Prod Type")

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[product_col]:
                    products.add(str(row[product_col]))

            # szukanie UN1845
            matching_products = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                cell_value = row[content_col]

                if cell_value and "UN1845" in str(cell_value).upper():
                    matching_products.add(str(cell_value))
                    self.is_dry_ice = True
                    break
                else:
                    self.is_dry_ice = False  

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row == "Pre-09":
                    self.is_t09 = True
                else:
                    self.is_t09 = False
            
            return list(products)
            
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił nieoczekiwany błąd: {e}")
            return []
    def toggle_batteries(self):
        if self.is_batteries == False:
            self.is_batteries = True
        else:    
            self.is_batteries = False
    def select_data(self):
        self.file_path = filedialog.askopenfilename(title="Wybierz plik Excel", filetypes=[("Pliki Excel", "*.xlsx")])
        if not self.file_path:
            return

        self.seal_number.delete(0, tk.END)
        self.seal_number.insert(0, self.file_path)

        self.products = self.read_excel(self.file_path)
        if not self.products:
            messagebox.showerror("Błąd", "Nie znaleziono danych w pliku Excel.")

    def check_if_data_available(self):
        if self.file_path:
            self.create_seal()
        else:
            messagebox.showerror("Błąd", "Nie wybrano pliku.")
            
    def create_seal(self):
        seal_parts = []
# sprawdza po zawartości czy są lody

        if self.is_dry_ice:
# baterie
            if self.bateries_lit_met.get() and self.bateries_lit_ion.get():
                seal_parts.append("*ICERLIRLM")
            elif self.bateries_lit_met.get():
                seal_parts.append("*ICERLM")
            elif self.bateries_lit_ion.get():
                seal_parts.append("*ICERLI")
            else:
                seal_parts.append("*ICE")

# soboty
        if self.saturday.get():
            seal_parts.append("DD6") 

        seal_parts.append("KTWGTU")
# drogówka
        if "W" in self.products and "H" in self.products:
            seal_parts.append("DDI")
        elif "H" in self.products:
            seal_parts.append("ESI")
        elif "W" in self.products:
            seal_parts.append("ESU")

# terminowki
        if "C" in self.products and "Q" in self.products:
            if "Y" in self.products or "T" in self.products or "K" in self.products:
                seal_parts.append("TMX")
            else:
                seal_parts.append("WMX")
        elif "Y" in self.products or "T" in self.products or "K" in self.products:
            if self.is_t09 == True:
                seal_parts.append("T09")
            else:
                seal_parts.append("T12")
        elif "C" in self.products:
            seal_parts.append("CMX")
        elif "Q" in self.products:
            seal_parts.append("WMX")
# zwykle paczki

        if "P" in self.products and "U" in self.products:
            seal_parts.append("MIP")
        elif "P" in self.products:
            seal_parts.append("WPX")
        elif "E" in self.products:
            seal_parts.append("ECX")
        elif "D" in self.products:
            seal_parts.append("DOX")
 


        match self.car_type.get():
            case 1:
                seal_parts.append("COY")
            case 2:
                seal_parts.append("NCY")
            case 3:
                seal_parts.append("CNY")
        seal_parts.append("ORGKRK")
        self.seal = "".join(seal_parts)


        if self.car_type.get() != 0:
            if len(self.seal) > 29:
                self.seal = self.seal[:-6]
            self.result_text.delete(0, tk.END)
            self.result_text.insert(0, self.seal)
            if self.type_of_batteries.get() == 1:
                messagebox.showwarning("Nie wybrano baterii"," Zaleca sie wybranie baterii. Jesli są załadowane.")
            if self.saturday.get() == False and self.dzis in(3,4) :
                messagebox.showwarning("Sobota?",f"Dzisiaj {self.name_of_day}. Sprawdź czy nie są załadowane paczki na sobotę!")
        else:
            messagebox.showerror("Błąd", "Nie wybrano typu auta.")
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()

